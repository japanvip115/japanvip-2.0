#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Sinh file Excel + CSV import Google Ads Editor cho Japan VIP (8 nhóm quảng cáo)."""
import csv, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

OUT_DIR = os.path.dirname(os.path.abspath(__file__)) + "/../google-ads"
os.makedirs(OUT_DIR, exist_ok=True)

CAMPAIGN = "Search - Gia dung Nhat (Ban hang)"

# ---- AD GROUPS & LANDING PAGES ----
AG1 = "Tu lanh Hitachi"
AG2 = "Bep tu Nhat"
AG3 = "Noi com dien Nhat"
AG4 = "May giat Nhat"
AG5 = "May loc khong khi Nhat"
AG6 = "May hut bui Dyson"
AG7 = "May rua bat Nhat"
AG8 = "May loc nuoc Nhat"

BASE = "https://japanvip.vn"
URL1 = BASE + "/danh-muc/tu-lanh-hitachi"
URL1_SP = BASE + "/tu-lanh-hitachi-r-wxc74x-x"  # landing page SP cụ thể (intent cao)
URL2 = BASE + "/danh-muc/bep-tu"
URL3 = BASE + "/danh-muc/noi-com-dien"
URL4 = BASE + "/danh-muc/may-giat"
URL5 = BASE + "/danh-muc/may-loc-khi"
URL6 = BASE + "/danh-muc/may-hut-bui"
URL7 = BASE + "/danh-muc/may-rua-bat"
URL8 = BASE + "/danh-muc/may-loc-nuoc"

# ---- KEYWORDS: (ad_group, keyword, match, maxcpc VND) ----
keywords = [
    # ===== AG1: Tủ lạnh Hitachi =====
    (AG1, "tủ lạnh hitachi nội địa nhật", "Exact", 12000),
    (AG1, "tủ lạnh hitachi r-wxc74x-x", "Exact", 15000),
    (AG1, "tủ lạnh hitachi r-wxc74x-x", "Phrase", 12000),
    (AG1, "tủ lạnh hitachi nhật bản", "Phrase", 10000),
    (AG1, "tủ lạnh hitachi 6 cánh nội địa nhật", "Phrase", 10000),
    (AG1, "tủ lạnh hitachi 735l", "Phrase", 9000),
    (AG1, "tủ lạnh hitachi hút chân không", "Phrase", 8000),
    (AG1, "mua tủ lạnh hitachi nội địa nhật", "Phrase", 8000),
    (AG1, "tủ lạnh hitachi r-wxc74x-x giá bao nhiêu", "Phrase", 10000),
    (AG1, "giá tủ lạnh hitachi nội địa nhật", "Phrase", 9000),
    (AG1, "tủ lạnh nhật 700 lít", "Phrase", 7000),
    (AG1, "tủ lạnh nội địa nhật cao cấp", "Phrase", 7000),
    (AG1, "tủ lạnh hitachi chính hãng", "Phrase", 8000),
    # Thêm từ khoá theo format đối thủ: [dung tích]L [số cánh] cánh (congnghenhat.com)
    (AG1, "tủ lạnh hitachi 735l 6 cánh", "Exact", 11000),
    (AG1, "tủ lạnh 6 cánh nội địa nhật", "Exact", 9000),
    (AG1, "tủ lạnh nhật 6 cửa cao cấp", "Phrase", 7000),
    (AG1, "tủ lạnh hitachi photocatalyst", "Phrase", 7000),
    (AG1, "tủ lạnh frost recycling nhật bản", "Phrase", 6000),

    # ===== AG2: Bếp từ Nhật =====
    (AG2, "bếp từ nội địa nhật", "Exact", 6000),
    (AG2, "bếp từ nhật bản", "Phrase", 5500),
    (AG2, "bếp từ panasonic nội địa nhật", "Phrase", 6000),
    (AG2, "bếp từ hitachi nội địa nhật", "Phrase", 6000),
    (AG2, "bếp từ mitsubishi nhật", "Phrase", 5500),
    (AG2, "mua bếp từ nội địa nhật", "Phrase", 5500),
    (AG2, "bếp từ đơn nội địa nhật", "Phrase", 5000),
    (AG2, "bếp từ 3 vùng nấu nhật", "Phrase", 5000),
    (AG2, "giá bếp từ nội địa nhật", "Phrase", 5000),
    (AG2, "bếp từ panasonic kz-l32as", "Exact", 7000),
    (AG2, "bếp từ hitachi ht-n8ktwf", "Exact", 7000),
    (AG2, "bếp từ nội địa nhật chính hãng", "Phrase", 5000),

    # ===== AG3: Nồi cơm điện Nhật =====
    (AG3, "nồi cơm điện nội địa nhật", "Exact", 5500),
    (AG3, "nồi cơm cao tần nhật", "Phrase", 5500),
    (AG3, "nồi cơm zojirushi nội địa nhật", "Phrase", 6000),
    (AG3, "nồi cơm tiger nhật bản", "Phrase", 5500),
    (AG3, "nồi cơm panasonic nội địa nhật", "Phrase", 5500),
    (AG3, "nồi cơm áp suất cao tần nhật", "Phrase", 5000),
    (AG3, "mua nồi cơm nhật nội địa", "Phrase", 5000),
    (AG3, "giá nồi cơm zojirushi nội địa nhật", "Phrase", 5500),
    (AG3, "nồi cơm ih áp suất nhật", "Phrase", 5000),
    (AG3, "nồi cơm nội địa nhật chính hãng", "Phrase", 4500),

    # ===== AG4: Máy giặt Nhật =====
    (AG4, "máy giặt nội địa nhật", "Exact", 6000),
    (AG4, "máy giặt nhật bản", "Phrase", 5500),
    (AG4, "máy giặt hitachi nội địa nhật", "Phrase", 6500),
    (AG4, "máy giặt panasonic nội địa nhật", "Phrase", 6000),
    (AG4, "máy giặt toshiba nội địa nhật", "Phrase", 6000),
    (AG4, "máy giặt nhật lồng đứng", "Phrase", 5500),
    (AG4, "máy giặt nhật lồng ngang", "Phrase", 5500),
    (AG4, "mua máy giặt nội địa nhật", "Phrase", 5500),
    (AG4, "giá máy giặt nội địa nhật", "Phrase", 5500),
    (AG4, "máy giặt nhật 12kg nội địa", "Phrase", 5000),
    (AG4, "máy giặt sấy khô nhật nội địa", "Phrase", 5000),
    (AG4, "máy giặt hitachi bd-nx120fr", "Exact", 7000),

    # ===== AG5: Máy lọc không khí Nhật =====
    (AG5, "máy lọc không khí nội địa nhật", "Exact", 5000),
    (AG5, "máy lọc khí nhật bản", "Phrase", 4500),
    (AG5, "máy lọc không khí panasonic nhật", "Phrase", 5000),
    (AG5, "máy lọc không khí sharp nhật", "Phrase", 5000),
    (AG5, "máy lọc không khí daikin nhật", "Phrase", 5000),
    (AG5, "mua máy lọc không khí nội địa nhật", "Phrase", 4500),
    (AG5, "máy lọc không khí nhật diệt khuẩn", "Phrase", 4500),
    (AG5, "máy lọc khí nhật tạo ion", "Phrase", 4000),
    (AG5, "giá máy lọc không khí nội địa nhật", "Phrase", 4500),
    (AG5, "máy lọc không khí nhật 40m2", "Phrase", 4000),

    # ===== AG6: Máy hút bụi Dyson =====
    (AG6, "máy hút bụi dyson nội địa nhật", "Exact", 8000),
    (AG6, "dyson nội địa nhật", "Phrase", 7500),
    (AG6, "máy hút bụi dyson nhật bản", "Phrase", 7000),
    (AG6, "dyson v12 nhật", "Phrase", 8000),
    (AG6, "dyson v10 nội địa nhật", "Phrase", 7500),
    (AG6, "mua dyson nội địa nhật", "Phrase", 7000),
    (AG6, "giá dyson nội địa nhật", "Phrase", 7000),
    (AG6, "máy hút bụi không dây nhật cao cấp", "Phrase", 6000),
    (AG6, "dyson chính hãng nhật bản", "Phrase", 7000),
    (AG6, "máy hút bụi dyson v15 nhật", "Exact", 9000),

    # ===== AG7: Máy rửa bát Nhật =====
    (AG7, "máy rửa bát nội địa nhật", "Exact", 6000),
    (AG7, "máy rửa bát nhật bản", "Phrase", 5500),
    (AG7, "máy rửa bát panasonic nội địa nhật", "Phrase", 6000),
    (AG7, "máy rửa bát panasonic np-tz500", "Exact", 7000),
    (AG7, "máy rửa bát panasonic np-tz300", "Exact", 6500),
    (AG7, "mua máy rửa bát nhật", "Phrase", 5500),
    (AG7, "giá máy rửa bát nội địa nhật", "Phrase", 5500),
    (AG7, "máy rửa bát để bàn nhật", "Phrase", 5000),
    (AG7, "máy rửa bát nhật tiết kiệm nước", "Phrase", 4500),

    # ===== AG8: Máy lọc nước Nhật =====
    (AG8, "máy lọc nước nội địa nhật", "Exact", 5500),
    (AG8, "máy lọc nước panasonic nhật", "Phrase", 6000),
    (AG8, "máy lọc nước ion kiềm nhật", "Phrase", 5500),
    (AG8, "máy lọc nước panasonic tk-as48", "Exact", 7000),
    (AG8, "mua máy lọc nước nội địa nhật", "Phrase", 5000),
    (AG8, "giá máy lọc nước ion kiềm nhật", "Phrase", 5000),
    (AG8, "máy lọc nước nhật gắn vòi", "Phrase", 4500),
    (AG8, "máy lọc nước điện giải nhật bản", "Phrase", 5000),
    (AG8, "máy lọc nước nhật bản chính hãng", "Phrase", 4500),
]

# ---- NEGATIVES (campaign-level, Phrase match) ----
negatives = [
    "cũ", "bãi", "hàng bãi", "thanh lý", "qua sử dụng", "second hand",
    "đã qua sử dụng", "sửa", "sửa chữa", "lỗi", "review", "đánh giá",
    "có tốt không", "là gì", "cách dùng", "hướng dẫn sử dụng", "thông số kỹ thuật",
    "tuyển dụng", "việc làm", "nội địa trung", "trung quốc", "hàng tàu",
    "mini", "thuê", "cho thuê", "fake", "nhái", "replica",
    "giá rẻ nhất thị trường", "siêu rẻ", "0 đồng", "miễn phí",
]

# ---- RSA ADS: (ad_group, final_url, path1, path2, [headlines 15], [descriptions 4]) ----
ads = [
    # ===== AG1: Tủ lạnh Hitachi — Ad 1 (features) =====
    (AG1, URL1, "tu-lanh", "hitachi",
     ["Tủ Lạnh Hitachi Nội Địa Nhật",
      "Hitachi Hút Chân Không 735L",
      "6 Cánh Cao Cấp Mới 100%",
      "Bảo Hành Chính Hãng 2 Năm",
      "Giao Lắp Tận Nơi Miễn Phí",
      "Đổi Trả 7 Ngày Không Lý Do",
      "Tủ Lạnh Hitachi Chính Hãng",
      "Japan VIP Uy Tín 10 Năm",
      "Gọi 09.2729.8888 Tư Vấn",
      "Trả Góp 0% Duyệt Nhanh",
      "Hàng Nhật Mới 100% Chính Hãng",
      "Giữ Thực Phẩm Tươi Lâu Hơn",
      "Hitachi 735L 6 Cánh Sẵn Hàng",
      "Showroom HN Xem Thực Tế",
      "Số 1 Hàng Nhật Tại Việt Nam"],
     ["Tủ lạnh Hitachi nội địa Nhật mới 100%. Bảo hành 2 năm, giao lắp miễn phí toàn quốc.",
      "Công nghệ hút chân không giữ thực phẩm tươi lâu hơn 40%. Trả góp 0%, đổi trả 7 ngày.",
      "Đủ model từ 400L–735L, 4–6 cánh. Xem thực tế tại showroom HN hoặc gọi 09.2729.8888.",
      "Japan VIP – nhập khẩu trực tiếp, không qua trung gian. Giá tốt, có tem nhập khẩu đầy đủ."]),

    # ===== AG1: Tủ lạnh Hitachi — Ad 2 (giá/ưu đãi) =====
    (AG1, URL1, "tu-lanh", "gia-tot",
     ["Tủ Lạnh Hitachi Giá Tốt Nhất",
      "Ưu Đãi Tủ Lạnh Nhật Hôm Nay",
      "Trả Góp 0% Giao Hàng Ngay",
      "Tủ Lạnh 735L Nội Địa Nhật",
      "Hàng Có Sẵn Giao Trong Ngày",
      "Bảo Hành Dài Chính Hãng",
      "Mua Trực Tiếp Giá Gốc",
      "Tủ Lạnh Hitachi 6 Cửa Cao Cấp",
      "Japan VIP Nhập Khẩu Trực Tiếp",
      "Gọi 09.2729.8888 Hỏi Giá",
      "Đổi Trả 7 Ngày Dễ Dàng",
      "Tư Vấn Miễn Phí 24/7",
      "Giao Lắp Miễn Phí HN & HCM",
      "Xuất VAT Đầy Đủ Theo Yêu Cầu",
      "Số 1 Hàng Nhật Tại VN"],
     ["Tủ lạnh Hitachi nội địa Nhật giá tốt nhất. Trả góp 0%, giao lắp tận nơi, không phí ẩn.",
      "Hàng mới 100%, bảo hành 2 năm. Mua trực tiếp Japan VIP – không qua trung gian.",
      "Ngăn hút chân không, tiết kiệm điện, công nghệ AI điều chỉnh nhiệt. Tư vấn 09.2729.8888.",
      "Có sẵn hàng, giao ngay nội thành. Đổi trả 7 ngày. Số 1 gia dụng nội địa Nhật tại VN."]),

    # ===== AG1: Tủ lạnh Hitachi — Ad 3 (model cụ thể R-WXC74X-X, landing SP) =====
    (AG1, URL1_SP, "tu-lanh", "r-wxc74x-x",
     ["Tủ Lạnh Hitachi R-WXC74X-X",
      "735L 6 Cánh Mới 100% Sẵn Hàng",
      "Hút Chân Không Photocatalyst",
      "Bảo Hành 2 Năm Tại Nhà",
      "Giao Lắp Miễn Phí HN & HCM",
      "Đổi Trả 7 Ngày Không Lý Do",
      "Hitachi R-WXC74X-X Chính Hãng",
      "Japan VIP Nhập Khẩu Trực Tiếp",
      "Gọi 09.2729.8888 Hỏi Giá",
      "Trả Góp 0% Duyệt Nhanh",
      "Frost Recycling Tiết Kiệm Điện",
      "Cửa Điện Trợ Lực Cao Cấp",
      "Hitachi 735L 6 Cánh Sẵn Hàng",
      "Xem Thực Tế Tại Showroom HN",
      "Số 1 Hàng Nhật Tại Việt Nam"],
     ["Tủ lạnh Hitachi R-WXC74X-X 735L 6 cánh nội địa Nhật. Sẵn hàng, giao lắp ngay nội thành.",
      "Hút chân không Photocatalyst giữ thực phẩm tươi lâu hơn 40%. Cửa điện trợ lực tiện dụng.",
      "Bảo hành 2 năm tại nhà. Trả góp 0%, đổi trả 7 ngày. Gọi 09.2729.8888 tư vấn ngay.",
      "Japan VIP nhập khẩu trực tiếp, không qua trung gian. Giá tốt, tem nhập khẩu đầy đủ."]),

    # ===== AG2: Bếp từ Nhật =====
    (AG2, URL2, "bep-tu", "noi-dia-nhat",
     ["Bếp Từ Nội Địa Nhật Chính Hãng",
      "Panasonic Hitachi Mitsubishi",
      "Bếp Từ Nhật Mới 100%",
      "Bảo Hành 2 Năm Chính Hãng",
      "Giao Lắp Tận Nơi Toàn Quốc",
      "12 Model Đang Có Sẵn",
      "Bếp Từ 3 Vùng Nấu Cao Cấp",
      "Tư Vấn Chọn Bếp Miễn Phí",
      "Japan VIP Uy Tín 10 Năm",
      "Gọi 09.2729.8888 Hỏi Giá",
      "Trả Góp 0% Duyệt Nhanh",
      "Đổi Trả 7 Ngày Không Lý Do",
      "Nấu Nhanh Tiết Kiệm Điện",
      "Showroom HN Xem Thực Tế",
      "Số 1 Hàng Nhật Tại VN"],
     ["Bếp từ nội địa Nhật Panasonic, Hitachi, Mitsubishi mới 100%. Bảo hành 2 năm chính hãng.",
      "Đủ loại đơn, đôi, 3 vùng nấu. Tiết kiệm điện, nấu nhanh, an toàn cho gia đình.",
      "Tư vấn chọn bếp miễn phí, giao lắp tận nơi. Trả góp 0%. Gọi 09.2729.8888.",
      "Japan VIP nhập khẩu trực tiếp, không qua trung gian. Giá tốt, có tem nhập khẩu đầy đủ."]),

    # ===== AG3: Nồi cơm điện Nhật =====
    (AG3, URL3, "noi-com", "noi-dia-nhat",
     ["Nồi Cơm Cao Tần Nội Địa Nhật",
      "Zojirushi Tiger Panasonic Nhật",
      "Cơm Dẻo Ngon Mỗi Bữa",
      "Công Nghệ IH Áp Suất Cao Cấp",
      "Bảo Hành Chính Hãng 2 Năm",
      "Quà Tặng Gia Đình Ý Nghĩa",
      "Nồi Cơm Nhật Mới 100%",
      "Japan VIP Uy Tín 10 Năm",
      "Gọi 09.2729.8888 Tư Vấn",
      "Trả Góp 0% Giao Toàn Quốc",
      "Đổi Trả 7 Ngày Dễ Dàng",
      "13 Model Đang Có Sẵn",
      "Giữ Ấm Cơm 12 Tiếng",
      "Tiết Kiệm Điện Công Nghệ Nhật",
      "Số 1 Hàng Nhật Tại VN"],
     ["Nồi cơm cao tần IH áp suất nội địa Nhật mới 100%. Zojirushi, Tiger, Panasonic chính hãng.",
      "Công nghệ IH áp suất cho cơm dẻo, ngon, giữ ấm 12 tiếng. Tiết kiệm điện vượt trội.",
      "Quà tặng ý nghĩa cho gia đình. Trả góp 0%, giao toàn quốc. Tư vấn miễn phí.",
      "Japan VIP nhập khẩu trực tiếp. Bảo hành 2 năm, đổi trả 7 ngày. Gọi 09.2729.8888."]),

    # ===== AG4: Máy giặt Nhật =====
    (AG4, URL4, "may-giat", "noi-dia-nhat",
     ["Máy Giặt Nhật Chính Hãng",
      "Hitachi Panasonic Toshiba Nhật",
      "Lồng Đứng Lồng Ngang Cao Cấp",
      "Giặt Sạch Tiết Kiệm Điện Nước",
      "Bảo Hành 2 Năm Chính Hãng",
      "Giao Lắp Tận Nơi Miễn Phí",
      "Máy Giặt Nhật Mới 100%",
      "12 Model Đang Có Sẵn",
      "Trả Góp 0% Duyệt Nhanh",
      "Japan VIP Uy Tín 10 Năm",
      "Gọi 09.2729.8888 Hỏi Giá",
      "Đổi Trả 7 Ngày Không Lý Do",
      "Công Nghệ Giặt AI Nhật Bản",
      "Showroom HN Xem Thực Tế",
      "Số 1 Hàng Nhật Tại VN"],
     ["Máy giặt nội địa Nhật Hitachi, Panasonic, Toshiba mới 100%. Bảo hành 2 năm chính hãng.",
      "Lồng đứng, lồng ngang 9–13kg. Tiết kiệm điện nước, công nghệ AI điều chỉnh chu trình giặt.",
      "Giao lắp tận nơi miễn phí, trả góp 0%. Đổi trả 7 ngày. Gọi 09.2729.8888.",
      "Japan VIP nhập khẩu trực tiếp, có tem nhập khẩu đầy đủ. Giá tốt, không qua trung gian."]),

    # ===== AG5: Máy lọc không khí Nhật =====
    (AG5, URL5, "may-loc-khi", "noi-dia-nhat",
     ["Máy Lọc Khí Nhật Chính Hãng",
      "Panasonic Sharp Daikin Nhật",
      "Diệt Khuẩn Ion Nano-e Nhật Bản",
      "Lọc PM2.5 Bụi Mịn Hiệu Quả",
      "Bảo Hành 2 Năm Chính Hãng",
      "Phù Hợp Phòng 20-60m2",
      "Máy Lọc Khí Nhật Mới 100%",
      "16 Model Đang Có Sẵn",
      "Trả Góp 0% Giao Toàn Quốc",
      "Japan VIP Uy Tín 10 Năm",
      "Gọi 09.2729.8888 Tư Vấn",
      "Đổi Trả 7 Ngày Dễ Dàng",
      "Vận Hành Êm Tiết Kiệm Điện",
      "Không Khí Sạch Cho Gia Đình",
      "Số 1 Hàng Nhật Tại VN"],
     ["Máy lọc không khí nội địa Nhật Panasonic, Sharp, Daikin mới 100%. Diệt khuẩn ion hiệu quả.",
      "Lọc PM2.5, bụi mịn, vi khuẩn. Phù hợp phòng 20–60m². Vận hành êm, tiết kiệm điện.",
      "16 model đang có sẵn. Trả góp 0%, giao toàn quốc. Tư vấn miễn phí 09.2729.8888.",
      "Japan VIP nhập khẩu trực tiếp. Bảo hành 2 năm, đổi trả 7 ngày. Không khí sạch."]),

    # ===== AG6: Máy hút bụi Dyson Nhật =====
    (AG6, URL6, "may-hut-bui", "dyson-nhat",
     ["Máy Hút Bụi Dyson Nội Địa Nhật",
      "Dyson V15 V12 V10 Chính Hãng",
      "Không Dây Hút Mạnh Gọn Nhẹ",
      "Lọc HEPA Diệt 99.97% Vi Khuẩn",
      "Bảo Hành 2 Năm Chính Hãng",
      "Dyson Nhật Mới 100%",
      "12 Model Đang Có Sẵn",
      "Trả Góp 0% Duyệt Nhanh",
      "Japan VIP Uy Tín 10 Năm",
      "Gọi 09.2729.8888 Hỏi Giá",
      "Đổi Trả 7 Ngày Không Lý Do",
      "Hút Cả Sàn Gỗ Và Thảm",
      "Phụ Kiện Đa Năng Kèm Theo",
      "Showroom HN Dùng Thử Trực Tiếp",
      "Số 1 Hàng Nhật Tại VN"],
     ["Máy hút bụi Dyson nội địa Nhật V10, V12, V15 mới 100%. Lọc HEPA, hút mạnh, pin lâu.",
      "Không dây, gọn nhẹ, vận hành êm. Diệt 99.97% vi khuẩn, bụi mịn. Phù hợp mọi sàn nhà.",
      "Trả góp 0%, giao toàn quốc. Bảo hành 2 năm chính hãng. Gọi 09.2729.8888 tư vấn.",
      "Japan VIP nhập khẩu trực tiếp từ Nhật. Giá tốt hơn xách tay, có tem nhập khẩu đầy đủ."]),

    # ===== AG7: Máy rửa bát Nhật =====
    (AG7, URL7, "may-rua-bat", "noi-dia-nhat",
     ["Máy Rửa Bát Nội Địa Nhật",
      "Panasonic NP-TZ500 Chính Hãng",
      "Rửa Sạch Sấy Khô 1 Lần",
      "Tiết Kiệm Nước Điện Hơn Tay",
      "Bảo Hành 2 Năm Chính Hãng",
      "Máy Rửa Bát Nhật Mới 100%",
      "Để Bàn Gắn Dưới Đều Có",
      "Trả Góp 0% Giao Lắp Miễn Phí",
      "Japan VIP Uy Tín 10 Năm",
      "Gọi 09.2729.8888 Tư Vấn",
      "Đổi Trả 7 Ngày Dễ Dàng",
      "6 Model Đang Có Sẵn",
      "Diệt Khuẩn 99.9% Bằng Nhiệt",
      "Showroom HN Xem Thực Tế",
      "Số 1 Hàng Nhật Tại VN"],
     ["Máy rửa bát Panasonic nội địa Nhật 100%. Rửa sạch, sấy khô, diệt khuẩn 99.9% bằng nhiệt.",
      "Tiết kiệm nước hơn rửa tay 7 lần. Để bàn hoặc gắn dưới bồn. Bảo hành 2 năm chính hãng.",
      "Trả góp 0%, giao lắp tận nơi miễn phí. Đổi trả 7 ngày. Gọi 09.2729.8888.",
      "Japan VIP nhập khẩu trực tiếp. Có tem nhập khẩu, xuất VAT đầy đủ. Số 1 hàng Nhật tại VN."]),

    # ===== AG8: Máy lọc nước Nhật =====
    (AG8, URL8, "may-loc-nuoc", "noi-dia-nhat",
     ["Máy Lọc Nước Ion Kiềm Nhật",
      "Panasonic TK-AS48 Chính Hãng",
      "Nước Kiềm Tốt Cho Sức Khỏe",
      "Lọc Trực Tiếp Từ Vòi Nước",
      "Bảo Hành 2 Năm Chính Hãng",
      "Máy Lọc Nước Nhật Mới 100%",
      "Gắn Dưới Bồn Tiện Lợi",
      "Trả Góp 0% Giao Lắp Miễn Phí",
      "Japan VIP Uy Tín 10 Năm",
      "Gọi 09.2729.8888 Tư Vấn",
      "Đổi Trả 7 Ngày Không Lý Do",
      "10 Model Đang Có Sẵn",
      "Diệt Khuẩn Lọc Sạch Clo",
      "Showroom HN Xem Thực Tế",
      "Số 1 Hàng Nhật Tại VN"],
     ["Máy lọc nước ion kiềm nội địa Nhật Panasonic mới 100%. Điện giải, lọc clo, diệt khuẩn.",
      "Nước kiềm pH 8.5–9.5 tốt cho tiêu hóa. Gắn trực tiếp vào vòi, không cần bình trữ.",
      "Trả góp 0%, giao lắp tận nơi miễn phí. Bảo hành 2 năm. Gọi 09.2729.8888.",
      "Japan VIP nhập khẩu trực tiếp từ Nhật. Có tem nhập khẩu, xuất VAT đầy đủ."]),
]

# ---- Kiểm tra giới hạn ký tự ----
warn = []
for ad in ads:
    for h in ad[4]:
        if len(h) > 30: warn.append(f"TIEU DE >30 ({len(h)}): {h}")
    for d in ad[5]:
        if len(d) > 90: warn.append(f"MO TA >90 ({len(d)}): {d}")
if warn:
    print("CANH BAO GIOI HAN:")
    for w in warn: print(" -", w)
else:
    print("OK: tat ca tieu de <=30, mo ta <=90 ky tu")

# ---- CSV: keywords ----
with open(f"{OUT_DIR}/1-keywords.csv", "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(["Campaign", "Ad Group", "Keyword", "Match Type", "Max CPC"])
    for ag, kw, mt, cpc in keywords:
        w.writerow([CAMPAIGN, ag, kw, mt, cpc])

# ---- CSV: negatives ----
with open(f"{OUT_DIR}/2-negative-keywords.csv", "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(["Campaign", "Keyword", "Match Type"])
    for n in negatives:
        w.writerow([CAMPAIGN, n, "Phrase"])

# ---- CSV: RSA ads ----
hcols = [f"Headline {i}" for i in range(1, 16)]
dcols = [f"Description {i}" for i in range(1, 5)]
with open(f"{OUT_DIR}/3-rsa-ads.csv", "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(["Campaign", "Ad Group"] + hcols + dcols + ["Final URL", "Path 1", "Path 2"])
    for ag, url, p1, p2, hs, ds in ads:
        w.writerow([CAMPAIGN, ag] + hs + ds + [url, p1, p2])

# ---- Excel tổng hợp ----
wb = openpyxl.Workbook()
hdr_fill = PatternFill("solid", fgColor="C41E3A")
hdr_font = Font(bold=True, color="FFFFFF")

def style_header(ws):
    for c in ws[1]:
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

ws = wb.active
ws.title = "1.Keywords"
ws.append(["Campaign", "Ad Group", "Keyword", "Match Type", "Max CPC (VND)"])
for ag, kw, mt, cpc in keywords:
    ws.append([CAMPAIGN, ag, kw, mt, cpc])
for col, wdt in zip("ABCDE", [32, 28, 42, 12, 14]):
    ws.column_dimensions[col].width = wdt
style_header(ws)

ws2 = wb.create_sheet("2.Negative Keywords")
ws2.append(["Campaign", "Keyword", "Match Type"])
for n in negatives:
    ws2.append([CAMPAIGN, n, "Phrase"])
for col, wdt in zip("ABC", [32, 32, 12]):
    ws2.column_dimensions[col].width = wdt
style_header(ws2)

ws3 = wb.create_sheet("3.RSA Ads")
ws3.append(["Campaign", "Ad Group"] + hcols + dcols + ["Final URL", "Path 1", "Path 2"])
for ag, url, p1, p2, hs, ds in ads:
    ws3.append([CAMPAIGN, ag] + hs + ds + [url, p1, p2])
style_header(ws3)
ws3.column_dimensions["A"].width = 32
ws3.column_dimensions["B"].width = 28

ws4 = wb.create_sheet("4.Cau truc Campaign")
structure = [
    ["CAMPAIGN: " + CAMPAIGN, "", "", ""],
    ["", "", "", ""],
    ["Ad Group", "Landing Page", "So tu khoa", "Ghi chu"],
    [AG1, URL1, str(sum(1 for x in keywords if x[0] == AG1)), "12 model tu lanh Hitachi, flagship R-WXC74X-X"],
    [AG2, URL2, str(sum(1 for x in keywords if x[0] == AG2)), "12 model bep tu Panasonic/Hitachi/Mitsubishi"],
    [AG3, URL3, str(sum(1 for x in keywords if x[0] == AG3)), "13 model noi com Zojirushi/Tiger/Panasonic"],
    [AG4, URL4, str(sum(1 for x in keywords if x[0] == AG4)), "12 model may giat Hitachi/Panasonic/Toshiba"],
    [AG5, URL5, str(sum(1 for x in keywords if x[0] == AG5)), "16 model may loc khi Panasonic/Sharp/Daikin"],
    [AG6, URL6, str(sum(1 for x in keywords if x[0] == AG6)), "12 model Dyson noi dia Nhat"],
    [AG7, URL7, str(sum(1 for x in keywords if x[0] == AG7)), "6 model may rua bat Panasonic"],
    [AG8, URL8, str(sum(1 for x in keywords if x[0] == AG8)), "10 model may loc nuoc Panasonic"],
]
for r in structure:
    ws4.append(r)
ws4["A1"].font = Font(bold=True, size=12, color="C41E3A")
for col, wdt in zip("ABCD", [30, 55, 14, 48]):
    ws4.column_dimensions[col].width = wdt
style_header(ws4)

ws5 = wb.create_sheet("5.Huong dan")
guide = [
    ["HUONG DAN NHAP VAO GOOGLE ADS EDITOR"],
    [""],
    ["B1. Tai & mo Google Ads Editor → dang nhap tai khoan → Get recent changes."],
    ["B2. Account menu → Import → From file → chon 1-keywords.csv → Process → Finish review."],
    ["B3. Tuong tu import 3-rsa-ads.csv (quang cao RSA)."],
    ["B4. Tu khoa phu dinh: Campaign → tab Negative keywords → dan cot Keyword tu 2-negative-keywords.csv (dat cap Campaign)."],
    ["B5. Kiem tra ngan sach Campaign (500k-1tr/ngay ~ 15-30tr/thang), khu vuc = Viet Nam, ngon ngu Tieng Viet."],
    ["B6. Post (day len tai khoan)."],
    [""],
    ["LUU Y QUAN TRONG:"],
    ["- 8 nhom quang cao, tong " + str(len(keywords)) + " tu khoa, " + str(len(ads)) + " RSA ads."],
    ["- Ngan sach goi y: uu tien AG1 (tu lanh Hitachi – gia tri cao) + AG6 (Dyson – CPC cao nhung intent manh)."],
    ["- Chay 3-7 ngay → xuat Search Terms report → them phu dinh moi, tang bid tu khoa chuyen doi tot."],
    ["- Gan theo doi chuyen doi GA4 (purchase) de toi uu theo don hang thuc te."],
    ["- Match Type: Exact = khop chinh xac, Phrase = khop cum tu."],
    ["- Max CPC chi la goi y khoi dau (VND) — dieu chinh theo thuc te dau thau."],
]
for r in guide:
    ws5.append(r)
ws5.column_dimensions["A"].width = 110
ws5["A1"].font = Font(bold=True, size=13, color="C41E3A")

xlsx_path = f"{OUT_DIR}/Google-Ads-Keywords-JapanVIP.xlsx"
wb.save(xlsx_path)
print("DA TAO:")
for p in ["Google-Ads-Keywords-JapanVIP.xlsx", "1-keywords.csv", "2-negative-keywords.csv", "3-rsa-ads.csv"]:
    print("  google-ads/" + p)
print(f"Tong: {len(keywords)} keywords, {len(negatives)} negatives, {len(ads)} RSA ads, 8 ad groups")
